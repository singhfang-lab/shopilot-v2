from __future__ import annotations

import asyncio
import json
import logging
import os
import secrets
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlmodel import Session, select

from .auth import get_merchant_for_user, require_admin
from .db import (
    BrandProfile, BrandReport,
    KBDocument, LLMConfig, Merchant, PlatformKBDocument, SystemPrompt,
    TestScenario,
    User, UserMerchant, get_db,
)
from . import rag, llm as llm_module
from .prompts import SYSTEM_PROMPT

router = APIRouter(prefix="/admin", tags=["admin"])

UPLOADS_DIR = Path.home() / "usb-assistant" / "uploads"
PLATFORM_UPLOADS_DIR = UPLOADS_DIR / "platform"
CLAUDE_API_BASE = "https://api.anthropic.com/v1"
CLAUDE_MODEL = "claude-sonnet-4-6"


# ── Stats ─────────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(admin=Depends(require_admin), db: Session = Depends(get_db)):
    user_count = len(db.exec(select(User)).all())
    merchant_count = len(db.exec(select(Merchant)).all())
    doc_count = len(db.exec(select(KBDocument)).all())
    platform_doc_count = len(db.exec(select(PlatformKBDocument)).all())
    active_prompt = db.exec(
        select(SystemPrompt).where(SystemPrompt.status == "active")
    ).first()
    return {
        "user_count": user_count,
        "merchant_count": merchant_count,
        "total_docs": doc_count,
        "platform_docs": platform_doc_count,
        "active_prompt_version": active_prompt.version if active_prompt else None,
        "active_prompt_label": active_prompt.label if active_prompt else "默认内置",
    }


# ── User management ───────────────────────────────────────────────────────────

@router.get("/users")
def list_users(admin=Depends(require_admin), db: Session = Depends(get_db)):
    users = db.exec(select(User)).all()
    result = []
    for u in users:
        merchant = get_merchant_for_user(u, db)
        result.append({
            "id": u.id,
            "email": u.email,
            "display_name": u.display_name,
            "role": u.role,
            "is_active": u.is_active,
            "created_at": u.created_at.isoformat(),
            "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
            "merchant": {"id": merchant.id, "name": merchant.name} if merchant else None,
        })
    return result


@router.post("/users/{user_id}/toggle-active")
def toggle_user_active(
    user_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot disable yourself")
    user.is_active = not user.is_active
    db.add(user)
    db.commit()
    return {"id": user.id, "is_active": user.is_active}


@router.post("/users/{user_id}/reset-password")
def reset_user_password(
    user_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    from .auth import _hash_password
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    tmp_password = secrets.token_urlsafe(8)
    user.password_hash = _hash_password(tmp_password)
    db.add(user)
    db.commit()
    return {"tmp_password": tmp_password}


# ── Platform KB ───────────────────────────────────────────────────────────────

@router.get("/platform-kb")
async def list_platform_kb(admin=Depends(require_admin), db: Session = Depends(get_db)):
    docs = db.exec(select(PlatformKBDocument)).all()
    sources = {s["source"]: s["chunk_count"] for s in await rag.list_sources_async(None)}
    return [
        {
            "id": d.id,
            "original_name": d.original_name,
            "filename": d.filename,
            "status": d.status,
            "chunk_count": sources.get(d.filename, d.chunk_count),
            "uploaded_at": d.uploaded_at.isoformat(),
            "indexed_at": d.indexed_at.isoformat() if d.indexed_at else None,
        }
        for d in docs
    ]


@router.post("/platform-kb/upload")
async def upload_platform_kb(
    file: UploadFile = File(...),
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    PLATFORM_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    bare_name = Path(file.filename).name if file.filename else "upload"
    safe_name = f"{secrets.token_hex(6)}_{bare_name}"
    save_path = PLATFORM_UPLOADS_DIR / safe_name
    content = await file.read()
    save_path.write_bytes(content)

    doc = PlatformKBDocument(
        filename=safe_name,
        original_name=file.filename or safe_name,
        status="processing",
        uploaded_by=admin.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    # Background indexing
    asyncio.create_task(_index_platform_doc(doc.id, save_path))
    return {"id": doc.id, "original_name": doc.original_name, "status": "processing"}


def _read_file_chunks(path: Path) -> list[str]:
    """Read a file and return text chunks."""
    suffix = path.suffix.lower()
    if suffix in (".txt", ".md"):
        text = path.read_text(encoding="utf-8", errors="ignore")
    elif suffix == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception:
            text = ""
    else:
        text = path.read_text(encoding="utf-8", errors="ignore")
    return rag._chunk_text(text) if text.strip() else []


async def _index_platform_doc(doc_id: int, path: Path):
    from .db import engine
    from sqlmodel import Session as S
    try:
        chunks = await asyncio.get_event_loop().run_in_executor(None, _read_file_chunks, path)
        if rag._use_pgvector():
            count = await rag._pg_add_async(chunks, source=path.name, shop_id=None)
        else:
            count = await asyncio.get_event_loop().run_in_executor(None, rag._chroma_add, chunks, path.name, None)
        with S(engine) as db:
            doc = db.get(PlatformKBDocument, doc_id)
            if doc:
                doc.status = "indexed"
                doc.chunk_count = count
                doc.indexed_at = datetime.now(timezone.utc)
                db.add(doc)
                db.commit()
    except Exception as e:
        logger.error("[platform-kb] indexing failed for doc_id=%s: %s", doc_id, e)
        with S(engine) as db:
            doc = db.get(PlatformKBDocument, doc_id)
            if doc:
                doc.status = "failed"
                db.add(doc)
                db.commit()


@router.delete("/platform-kb/{doc_id}")
async def delete_platform_kb(
    doc_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    doc = db.get(PlatformKBDocument, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if rag._use_pgvector():
        deleted = await rag._pg_delete_by_source_async(doc.filename, shop_id=None)
    else:
        deleted = rag.delete_by_source(doc.filename, shop_id=None)
    path = PLATFORM_UPLOADS_DIR / doc.filename
    if path.exists():
        path.unlink()
    db.delete(doc)
    db.commit()
    return {"ok": True, "chunks_removed": deleted}


class RenamePlatformKBRequest(BaseModel):
    new_name: str  # just the filename part, no path


@router.patch("/platform-kb/{doc_id}/rename")
async def rename_platform_kb(
    doc_id: int,
    req: RenamePlatformKBRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    doc = db.get(PlatformKBDocument, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    # Keep the folder prefix, replace only the filename part
    parts = doc.original_name.rsplit("/", 1)
    if len(parts) == 2:
        doc.original_name = f"{parts[0]}/{req.new_name}"
    else:
        doc.original_name = req.new_name
    db.add(doc)
    db.commit()
    return {"ok": True, "original_name": doc.original_name}


class CreateFolderRequest(BaseModel):
    path: str  # e.g. "新建文件夹" or "父目录/子目录"


@router.post("/platform-kb-folder")
async def create_folder_platform_kb(
    req: CreateFolderRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Create a virtual folder by inserting a .keep placeholder document."""
    folder_path = req.path.strip("/").strip()
    if not folder_path:
        raise HTTPException(status_code=400, detail="目录名不能为空")
    placeholder_name = f"{folder_path}/.keep"
    # Check if already exists
    existing = db.exec(
        select(PlatformKBDocument).where(PlatformKBDocument.original_name == placeholder_name)
    ).first()
    if existing:
        return {"ok": True, "id": existing.id, "original_name": existing.original_name}

    PLATFORM_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = f"{secrets.token_hex(6)}_.keep"
    save_path = PLATFORM_UPLOADS_DIR / safe_name
    save_path.write_bytes(b"")

    doc = PlatformKBDocument(
        filename=safe_name,
        original_name=placeholder_name,
        status="indexed",
        chunk_count=0,
        uploaded_by=admin.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return {"ok": True, "id": doc.id, "original_name": doc.original_name}


class RenameFolderRequest(BaseModel):
    old_prefix: str   # e.g. "新建文件夹/07_营销活动与增长"
    new_prefix: str   # e.g. "新建文件夹/07_营销推广与增长"


@router.post("/platform-kb/rename-folder")
async def rename_folder_platform_kb(
    req: RenameFolderRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    docs = db.exec(select(PlatformKBDocument)).all()
    updated = 0
    for doc in docs:
        if doc.original_name.startswith(req.old_prefix + "/") or doc.original_name == req.old_prefix:
            doc.original_name = req.new_prefix + doc.original_name[len(req.old_prefix):]
            db.add(doc)
            updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


class DeleteFolderRequest(BaseModel):
    prefix: str   # e.g. "新建文件夹/07_营销活动与增长"


@router.delete("/platform-kb-folder")
async def delete_folder_platform_kb(
    req: DeleteFolderRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    docs = db.exec(select(PlatformKBDocument)).all()
    to_delete = [
        d for d in docs
        if d.original_name.startswith(req.prefix + "/") or d.original_name == req.prefix
    ]
    total_chunks = 0
    for doc in to_delete:
        if rag._use_pgvector():
            deleted = await rag._pg_delete_by_source_async(doc.filename, shop_id=None)
        else:
            deleted = rag.delete_by_source(doc.filename, shop_id=None)
        total_chunks += deleted or 0
        path = PLATFORM_UPLOADS_DIR / doc.filename
        if path.exists():
            path.unlink()
        db.delete(doc)
    db.commit()
    return {"ok": True, "files_deleted": len(to_delete), "chunks_removed": total_chunks}


# ── Prompt management ─────────────────────────────────────────────────────────

@router.get("/prompts")
def list_prompts(
    prompt_type: str = "chat",
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    prompts = db.exec(
        select(SystemPrompt)
        .where(SystemPrompt.prompt_type == prompt_type)
        .order_by(SystemPrompt.version.desc())
    ).all()
    return [
        {
            "id": p.id,
            "version": p.version,
            "label": p.label,
            "status": p.status,
            "prompt_type": p.prompt_type,
            "created_at": p.created_at.isoformat(),
            "published_at": p.published_at.isoformat() if p.published_at else None,
            "content_preview": p.content[:120] + "..." if len(p.content) > 120 else p.content,
        }
        for p in prompts
    ]


@router.get("/prompts/active")
def get_active_prompt(admin=Depends(require_admin), db: Session = Depends(get_db)):
    p = db.exec(select(SystemPrompt).where(SystemPrompt.status == "active")).first()
    if not p:
        return {"content": SYSTEM_PROMPT, "version": None, "label": "默认内置", "status": "active"}
    return {"id": p.id, "version": p.version, "label": p.label, "status": p.status, "content": p.content}


@router.get("/prompts/{prompt_id}")
def get_prompt(prompt_id: int, admin=Depends(require_admin), db: Session = Depends(get_db)):
    p = db.get(SystemPrompt, prompt_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    return {
        "id": p.id, "version": p.version, "label": p.label,
        "status": p.status, "content": p.content,
        "test_result": p.test_result,
        "created_at": p.created_at.isoformat(),
        "published_at": p.published_at.isoformat() if p.published_at else None,
    }


class CreatePromptRequest(BaseModel):
    content: str
    label: str = ""
    prompt_type: str = "chat"


@router.post("/prompts")
def create_prompt(
    req: CreatePromptRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    latest = db.exec(
        select(SystemPrompt)
        .where(SystemPrompt.prompt_type == req.prompt_type)
        .order_by(SystemPrompt.version.desc())
    ).first()
    next_version = (latest.version + 1) if latest else 1
    p = SystemPrompt(
        version=next_version,
        content=req.content,
        label=req.label or f"V{next_version}",
        status="draft",
        prompt_type=req.prompt_type,
        created_by=admin.id,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"id": p.id, "version": p.version, "label": p.label, "status": p.status}


class UpdatePromptRequest(BaseModel):
    content: str
    label: str = ""


@router.patch("/prompts/{prompt_id}")
def update_prompt(
    prompt_id: int,
    req: UpdatePromptRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.get(SystemPrompt, prompt_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    if p.status not in ("draft", "testing"):
        raise HTTPException(status_code=400, detail="Only draft/testing prompts can be edited")
    p.content = req.content
    if req.label:
        p.label = req.label
    db.add(p)
    db.commit()
    return {"ok": True}


class AISuggestRequest(BaseModel):
    current_content: str
    instruction: str


@router.post("/prompts/ai-suggest")
async def ai_suggest_prompt(
    req: AISuggestRequest,
    admin=Depends(require_admin),
):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="Anthropic API key not configured")

    meta_prompt = f"""你是一位 AI Prompt 工程专家。请根据以下优化指令，对现有 System Prompt 进行改写。

优化指令：{req.instruction}

现有 Prompt：
{req.current_content}

请直接输出改写后的完整 Prompt 文本，不要加任何前缀说明。然后在最后另起一行输出：
---EXPLANATION---
（简短说明你做了哪些改动及原因，2-3句话即可）"""

    async with httpx.AsyncClient(timeout=60.0) as client:
        res = await client.post(
            f"{CLAUDE_API_BASE}/messages",
            json={"model": CLAUDE_MODEL, "max_tokens": 4096, "temperature": 0.4,
                  "messages": [{"role": "user", "content": meta_prompt}]},
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01"},
        )
    if res.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Claude error {res.status_code}")

    text = res.json()["content"][0]["text"]
    if "---EXPLANATION---" in text:
        parts = text.split("---EXPLANATION---", 1)
        return {"suggested_content": parts[0].strip(), "explanation": parts[1].strip()}
    return {"suggested_content": text.strip(), "explanation": ""}


class TestPromptRequest(BaseModel):
    test_message: str
    provider: str = "claude"
    model: str = ""


@router.post("/prompts/{prompt_id}/test")
async def test_prompt(
    prompt_id: int,
    req: TestPromptRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.get(SystemPrompt, prompt_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")

    # Look up the llm_config for the requested provider
    cfg = db.exec(
        select(LLMConfig).where(LLMConfig.provider == req.provider)
    ).first()

    provider = req.provider
    model = req.model or (cfg.model if cfg else CLAUDE_MODEL)
    api_key = (cfg.api_key if cfg and cfg.api_key else None)

    if provider == "claude":
        api_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=503, detail="Anthropic API key not configured")
        async with httpx.AsyncClient(timeout=60.0) as client:
            res = await client.post(
                f"{CLAUDE_API_BASE}/messages",
                json={"model": model, "max_tokens": 800, "system": p.content,
                      "messages": [{"role": "user", "content": req.test_message}]},
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01"},
            )
        if res.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Claude error {res.status_code}: {res.text[:200]}")
        reply = res.json()["content"][0]["text"]

    elif provider == "openai":
        api_key = api_key or os.environ.get("OPENAI_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=503, detail="OpenAI API key not configured")
        async with httpx.AsyncClient(timeout=60.0) as client:
            res = await client.post(
                "https://api.openai.com/v1/chat/completions",
                json={"model": model, "max_tokens": 800,
                      "messages": [{"role": "system", "content": p.content},
                                   {"role": "user", "content": req.test_message}]},
                headers={"Authorization": f"Bearer {api_key}"},
            )
        if res.status_code != 200:
            raise HTTPException(status_code=502, detail=f"OpenAI error {res.status_code}: {res.text[:200]}")
        reply = res.json()["choices"][0]["message"]["content"]

    elif provider == "gemini":
        api_key = api_key or os.environ.get("GEMINI_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=503, detail="Gemini API key not configured")
        async with httpx.AsyncClient(timeout=60.0) as client:
            res = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}",
                json={"system_instruction": {"parts": [{"text": p.content}]},
                      "contents": [{"parts": [{"text": req.test_message}]}]},
            )
        if res.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Gemini error {res.status_code}: {res.text[:200]}")
        reply = res.json()["candidates"][0]["content"]["parts"][0]["text"]

    elif provider == "ollama":
        base = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
        async with httpx.AsyncClient(timeout=120.0) as client:
            res = await client.post(
                f"{base}/api/chat",
                json={"model": model, "stream": False,
                      "messages": [{"role": "system", "content": p.content},
                                   {"role": "user", "content": req.test_message}]},
            )
        if res.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Ollama error {res.status_code}: {res.text[:200]}")
        reply = res.json()["message"]["content"]

    else:
        raise HTTPException(status_code=400, detail=f"Unsupported provider: {provider}")

    p.test_result = json.dumps({"message": req.test_message, "reply": reply,
                                "provider": provider, "model": model}, ensure_ascii=False)
    p.status = "testing"
    db.add(p)
    db.commit()
    return {"reply": reply, "provider": provider, "model": model}


@router.post("/prompts/{prompt_id}/publish")
def publish_prompt(
    prompt_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.get(SystemPrompt, prompt_id)
    if not p:
        raise HTTPException(status_code=404, detail="Not found")

    # Archive current active
    current_active = db.exec(select(SystemPrompt).where(SystemPrompt.status == "active")).all()
    for old in current_active:
        old.status = "archived"
        db.add(old)

    p.status = "active"
    p.published_at = datetime.now(timezone.utc)
    db.add(p)
    db.commit()
    return {"ok": True, "version": p.version}


@router.post("/prompts/{prompt_id}/rollback")
def rollback_prompt(
    prompt_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    return publish_prompt(prompt_id, admin, db)


# ── LLM config management ─────────────────────────────────────────────────────

@router.get("/llm-configs")
def list_llm_configs(admin=Depends(require_admin), db: Session = Depends(get_db)):
    configs = db.exec(select(LLMConfig)).all()
    env_keys = {
        "gemini": os.environ.get("GEMINI_API_KEY", ""),
        "claude": os.environ.get("ANTHROPIC_API_KEY", ""),
        "openai": os.environ.get("OPENAI_API_KEY", ""),
    }
    return [
        {
            "id": c.id,
            "provider": c.provider,
            "model": c.model or llm_module.DEFAULT_MODELS.get(c.provider, ""),
            "label": c.label,
            "is_active": c.is_active,
            "has_key": bool(c.api_key) or bool(env_keys.get(c.provider, "")),
            "updated_at": c.updated_at.isoformat(),
        }
        for c in configs
    ]


class LLMConfigUpdateRequest(BaseModel):
    api_key: str = ""
    label: str = ""
    model: str = ""


@router.patch("/llm-configs/{config_id}")
def update_llm_config(
    config_id: int,
    req: LLMConfigUpdateRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    cfg = db.get(LLMConfig, config_id)
    if not cfg:
        raise HTTPException(status_code=404, detail="Not found")
    if req.api_key:
        cfg.api_key = req.api_key
    if req.label:
        cfg.label = req.label
    if req.model:
        cfg.model = req.model
    cfg.updated_at = datetime.now(timezone.utc)
    db.add(cfg)
    db.commit()
    return {"ok": True}


@router.post("/llm-configs/{config_id}/toggle-enabled")
def toggle_llm_config_enabled(
    config_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Toggle whether this config appears in the frontend model selector."""
    cfg = db.get(LLMConfig, config_id)
    if not cfg:
        raise HTTPException(status_code=404, detail="Not found")
    cfg.is_active = not cfg.is_active
    cfg.updated_at = datetime.now(timezone.utc)
    db.add(cfg)
    db.commit()
    return {"ok": True, "is_active": cfg.is_active}


@router.delete("/llm-configs/{config_id}")
def delete_llm_config(
    config_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    cfg = db.get(LLMConfig, config_id)
    if not cfg:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(cfg)
    db.commit()
    return {"ok": True}


# ── Scenario Testing ──────────────────────────────────────────────────────────

SCENARIO_DATA_DIR = Path("/Users/singhfang/Downloads/问题库&模拟商家数据")

# In-memory task store: task_id → {status, progress, results, html, error}
_test_tasks: dict[str, dict] = {}


@router.get("/test-scenarios/data-files")
def list_scenario_data_files(admin=Depends(require_admin)):
    """Return xlsx/csv files available in the scenario DATA_DIR."""
    if not SCENARIO_DATA_DIR.exists():
        return []
    files = sorted(
        f.name for f in SCENARIO_DATA_DIR.iterdir()
        if f.is_file() and f.suffix.lower() in (".xlsx", ".xls", ".csv")
    )
    return files


@router.post("/test-scenarios/data-files/upload")
async def upload_scenario_data_file(
    file: UploadFile = File(...),
    admin=Depends(require_admin),
):
    """Upload a new test data file into the scenario DATA_DIR."""
    SCENARIO_DATA_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename).name
    dest = SCENARIO_DATA_DIR / safe_name
    content = await file.read()
    dest.write_bytes(content)
    return {"filename": safe_name}



class RunScenariosRequest(BaseModel):
    cases: list[str] = []
    no_judge: bool = False


@router.post("/test-scenarios/run")
def run_test_scenarios(req: RunScenariosRequest, admin=Depends(require_admin)):
    from .test_scenarios import SCENE_META, run_scenario, build_html_report

    all_ids = list(SCENE_META.keys())
    run_ids = [c.upper() for c in req.cases if c.upper() in all_ids] or all_ids

    task_id = uuid.uuid4().hex
    _test_tasks[task_id] = {
        "status": "running",
        "progress": {"done": 0, "total": len(run_ids)},
        "results": [],
        "html": None,
        "error": None,
    }

    CONCURRENCY = 5  # 最多同时跑 5 个场景

    def _worker():
        import time
        from .test_scenarios import ScenarioResult
        results_map: dict[str, object] = {}
        t0 = time.time()
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        async def _run_one(sid: str):
            try:
                r = await run_scenario(sid, no_judge=req.no_judge)
            except Exception as e:
                r = ScenarioResult(
                    sid=sid,
                    name=SCENE_META[sid]["name"],
                    business_type=SCENE_META[sid]["business_type"],
                    error=str(e),
                )
            results_map[sid] = r
            _test_tasks[task_id]["progress"]["done"] = len(results_map)
            # preserve original order for results
            ordered = [results_map[s] for s in run_ids if s in results_map]
            _test_tasks[task_id]["results"] = _serialise_results(ordered)

        async def _run_all():
            sem = asyncio.Semaphore(CONCURRENCY)
            async def _guarded(sid):
                async with sem:
                    await _run_one(sid)
            await asyncio.gather(*[_guarded(sid) for sid in run_ids])

        try:
            loop.run_until_complete(_run_all())
            ordered = [results_map[s] for s in run_ids if s in results_map]
            total_ms = int((time.time() - t0) * 1000)
            html = build_html_report(ordered, total_ms)
            _test_tasks[task_id]["status"] = "done"
            _test_tasks[task_id]["html"] = html
            _test_tasks[task_id]["results"] = _serialise_results(ordered)
        except Exception as e:
            _test_tasks[task_id]["status"] = "error"
            _test_tasks[task_id]["error"] = str(e)
        finally:
            loop.close()

    threading.Thread(target=_worker, daemon=True).start()
    return {"task_id": task_id, "total": len(run_ids)}


def _serialise_results(results) -> list[dict]:
    out = []
    for sc in results:
        turns = []
        for t in sc.turns:
            s = t.score
            turns.append({
                "turn_no": t.turn_no,
                "user_msg": t.user_msg,
                "ai_response": t.ai_response,
                "has_card": t.has_card,
                "has_chart": t.has_chart,
                "has_tool_call": t.has_tool_call,
                "chars": t.chars,
                "latency_ms": t.latency_ms,
                "score": {
                    "overall": s.overall,
                    "accuracy": s.accuracy,
                    "actionability": s.actionability,
                    "completeness": s.completeness,
                    "clarity": s.clarity,
                    "relevance": s.relevance,
                    "data_usage": s.data_usage,
                    "strengths": s.strengths,
                    "weaknesses": s.weaknesses,
                    "red_flags": s.red_flags,
                    "summary": s.summary,
                    "error": s.error,
                },
            })
        sc_scores = [t.score.overall for t in sc.turns if t.score and not t.score.error and t.score.overall > 0]
        avg = round(sum(sc_scores) / len(sc_scores), 2) if sc_scores else 0
        out.append({
            "sid": sc.sid,
            "name": sc.name,
            "business_type": sc.business_type,
            "error": sc.error,
            "avg_score": avg,
            "red_flag_hits": sc.red_flag_hits,
            "must_hit": sc.must_hit,
            "must_hit_pass": sc.must_hit_pass,
            "turns": turns,
        })
    return out


@router.get("/test-scenarios/result/{task_id}")
def get_test_result(task_id: str, admin=Depends(require_admin)):
    task = _test_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


# ── Brand library ─────────────────────────────────────────────────────────────

@router.get("/brands")
def list_brands(
    q: str = "",
    business_type: str = "",
    skip: int = 0,
    limit: int = 100,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    stmt = select(BrandProfile)
    if q:
        stmt = stmt.where(BrandProfile.name.ilike(f"%{q}%"))
    if business_type:
        stmt = stmt.where(BrandProfile.business_type == business_type)
    stmt = stmt.order_by(BrandProfile.business_type, BrandProfile.name).offset(skip).limit(limit)
    brands = db.exec(stmt).all()

    # Check which have reports
    brand_ids = [b.id for b in brands]
    reports = db.exec(
        select(BrandReport.brand_id, BrandReport.generated_at, BrandReport.model_used)
        .where(BrandReport.brand_id.in_(brand_ids))
    ).all()
    report_map = {r.brand_id: r for r in reports}

    return [
        {
            "id": b.id,
            "name": b.name,
            "business_type": b.business_type,
            "region": b.region,
            "store_count": b.store_count,
            "avg_price_amap": b.avg_price_amap,
            "avg_rating_amap": b.avg_rating_amap,
            "amap_sample_size": b.amap_sample_size,
            "price_tier": b.price_tier,
            "is_active": b.is_active,
            "has_report": b.id in report_map,
            "report_generated_at": report_map[b.id].generated_at.isoformat() if b.id in report_map else None,
            "has_profile_json": bool(b.profile_json),
        }
        for b in brands
    ]


@router.get("/brands/{brand_id}")
def get_brand(brand_id: int, admin=Depends(require_admin), db: Session = Depends(get_db)):
    b = db.get(BrandProfile, brand_id)
    if not b:
        raise HTTPException(status_code=404, detail="Brand not found")

    rpt = db.exec(
        select(BrandReport)
        .where(BrandReport.brand_id == brand_id)
        .order_by(BrandReport.generated_at.desc())
    ).first()

    return {
        "id": b.id,
        "name": b.name,
        "business_type": b.business_type,
        "region": b.region,
        "founded_year": b.founded_year,
        "store_count": b.store_count,
        "store_count_year": b.store_count_year,
        "headquarters": b.headquarters,
        "revenue": b.revenue,
        "revenue_year": b.revenue_year,
        "avg_price_min": b.avg_price_min,
        "avg_price_max": b.avg_price_max,
        "price_band": b.price_band,
        "avg_price_amap": b.avg_price_amap,
        "avg_rating_amap": b.avg_rating_amap,
        "amap_sample_size": b.amap_sample_size,
        "price_tier": b.price_tier,
        "quality_perception": b.quality_perception,
        "delivery_rate": b.delivery_rate,
        "main_competitors": b.main_competitors,
        "user_tags": b.user_tags,
        "description": b.description,
        "aliases": b.aliases,
        "markets": b.markets,
        "expansion": b.expansion,
        "is_active": b.is_active,
        "profile_json": b.profile_json,
        "report": {
            "id": rpt.id,
            "report_json": rpt.report_json,
            "prompt_template": rpt.prompt_template,
            "generated_at": rpt.generated_at.isoformat(),
            "model_used": rpt.model_used,
        } if rpt else None,
    }


class BrandUpdateRequest(BaseModel):
    description: str | None = None
    main_competitors: str | None = None   # JSON array string
    user_tags: str | None = None          # JSON array string
    store_count: int | None = None
    founded_year: int | None = None
    headquarters: str | None = None
    price_tier: str | None = None
    quality_perception: str | None = None
    expansion: str | None = None
    is_active: bool | None = None


@router.patch("/brands/{brand_id}")
def update_brand(
    brand_id: int,
    req: BrandUpdateRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    b = db.get(BrandProfile, brand_id)
    if not b:
        raise HTTPException(status_code=404, detail="Brand not found")
    for field, val in req.model_dump(exclude_none=True).items():
        setattr(b, field, val)
    b.updated_at = datetime.now(timezone.utc)
    db.add(b)
    db.commit()
    return {"ok": True}


@router.post("/brands/{brand_id}/regen-report")
async def regen_brand_report(
    brand_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    b = db.get(BrandProfile, brand_id)
    if not b:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Import here to avoid circular at module load
    from .scripts.seed_cn_brands import _generate_report

    result = await asyncio.get_event_loop().run_in_executor(None, _generate_report, brand_id)
    if not result:
        raise HTTPException(status_code=500, detail="Report generation failed")

    report_json, prompt_template = result
    rpt = BrandReport(
        brand_id=brand_id,
        report_json=report_json,
        prompt_template=prompt_template,
        model_used="claude-sonnet-4-6",
        generated_at=datetime.now(timezone.utc),
    )
    db.add(rpt)
    db.commit()
    db.refresh(rpt)
    return {"ok": True, "report_id": rpt.id, "generated_at": rpt.generated_at.isoformat()}


# ── Brand prefill (AI) ───────────────────────────────────────────────────────

class PrefillBrandRequest(BaseModel):
    name: str
    business_type: str = ""


@router.post("/brands/prefill")
async def prefill_brand(req: PrefillBrandRequest, admin=Depends(require_admin)):
    """Use Claude Haiku to prefill brand fields from training knowledge."""
    import anthropic as _anthropic
    import re as _re

    client = _anthropic.Anthropic()
    prompt = f"""根据你对「{req.name}」{f'（{req.business_type}品牌）' if req.business_type else ''}的了解，填写以下 JSON 字段。无法确认的填 null，price_tier 和 quality_perception 必须给值。

{{
  "business_type": "品类，如大众咖啡连锁/火锅连锁/快餐连锁等",
  "founded_year": 数字或null,
  "store_count": 估算门店数或null,
  "headquarters": "总部城市，格式如北京市",
  "avg_price_min": 客单价下限元或null,
  "avg_price_max": 客单价上限元或null,
  "price_tier": "低价/中价/高价",
  "quality_perception": "低/中/高",
  "main_competitors": ["竞品1", "竞品2", "竞品3"],
  "user_tags": ["用户认知标签1", "标签2", "标签3"],
  "description": "品牌简介2-3句话"
}}

只返回 JSON，不要任何说明。"""

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        text = msg.content[0].text
        m = _re.search(r'\{[\s\S]+\}', text)
        if m:
            import json as _json
            return _json.loads(m.group())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    raise HTTPException(status_code=500, detail="AI 预填失败")


# ── Brand create + init ───────────────────────────────────────────────────────

class CreateBrandRequest(BaseModel):
    name: str
    business_type: str
    region: str = "cn"
    description: str = ""
    headquarters: str = ""
    founded_year: Optional[int] = None
    store_count: Optional[int] = None
    price_tier: str = ""
    quality_perception: str = ""
    main_competitors: str = "[]"
    user_tags: str = "[]"


@router.post("/brands", status_code=201)
def create_brand(
    req: CreateBrandRequest,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    existing = db.exec(select(BrandProfile).where(BrandProfile.name == req.name)).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"品牌「{req.name}」已存在 (id={existing.id})")
    b = BrandProfile(
        name=req.name,
        business_type=req.business_type,
        region=req.region,
        description=req.description,
        headquarters=req.headquarters,
        founded_year=req.founded_year,
        store_count=req.store_count,
        price_tier=req.price_tier,
        quality_perception=req.quality_perception,
        main_competitors=req.main_competitors,
        user_tags=req.user_tags,
        is_active=True,
    )
    db.add(b)
    db.commit()
    db.refresh(b)
    return {"id": b.id, "name": b.name}


# In-memory init status store
_init_tasks: dict[int, dict] = {}   # brand_id -> {status, step, error}


@router.post("/brands/{brand_id}/init")
def init_brand(brand_id: int, admin=Depends(require_admin), db: Session = Depends(get_db)):
    b = db.get(BrandProfile, brand_id)
    if not b:
        raise HTTPException(status_code=404, detail="Brand not found")
    if _init_tasks.get(brand_id, {}).get("status") == "running":
        return {"status": "running"}

    _init_tasks[brand_id] = {"status": "running", "step": "starting", "error": None}

    def _run():
        import asyncio as _asyncio
        from .scripts.seed_cn_brands import _amap_sample, _generate_report, _generate_profile_json

        loop = _asyncio.new_event_loop()
        try:
            # Step 1: Amap sampling
            _init_tasks[brand_id]["step"] = "amap"
            with Session(db.bind) as s:
                brand = s.get(BrandProfile, brand_id)
            amap = loop.run_until_complete(_amap_sample(brand.name, brand.headquarters or ""))
            if amap:
                with Session(db.bind) as s:
                    brand = s.get(BrandProfile, brand_id)
                    brand.avg_price_amap = amap.get("avg_price")
                    brand.avg_rating_amap = amap.get("avg_rating")
                    brand.delivery_rate = amap.get("delivery_rate")
                    brand.amap_sample_size = amap.get("sample_size", 0)
                    s.add(brand)
                    s.commit()

            # Step 2: Generate report
            _init_tasks[brand_id]["step"] = "report"
            result = _generate_report(brand_id)
            if result:
                report_json, prompt_template = result
                with Session(db.bind) as s:
                    s.add(BrandReport(
                        brand_id=brand_id,
                        report_json=report_json,
                        prompt_template=prompt_template,
                        model_used="claude-sonnet-4-6",
                        generated_at=datetime.now(timezone.utc),
                    ))
                    s.commit()

            # Step 3: Generate profile_json
            _init_tasks[brand_id]["step"] = "profile"
            profile_json = _generate_profile_json(brand_id)
            if profile_json:
                with Session(db.bind) as s:
                    brand = s.get(BrandProfile, brand_id)
                    brand.profile_json = profile_json
                    s.add(brand)
                    s.commit()

            _init_tasks[brand_id] = {"status": "done", "step": "done", "error": None}

        except Exception as e:
            _init_tasks[brand_id] = {"status": "error", "step": _init_tasks[brand_id].get("step", ""), "error": str(e)}
        finally:
            loop.close()

    threading.Thread(target=_run, daemon=True).start()
    return {"status": "running"}


@router.get("/brands/{brand_id}/init-status")
def get_init_status(brand_id: int, admin=Depends(require_admin)):
    task = _init_tasks.get(brand_id)
    if not task:
        return {"status": "idle"}
    return task


# ── Test Scenario CRUD ────────────────────────────────────────────────────────

class CreateScenarioRequest(BaseModel):
    sid: str
    name: str
    business_type: str = ""
    shop_name: str = ""
    shop_category: str = ""
    shop_address: str = ""
    csv_name: str = ""
    q1: str = ""
    q2: str = ""
    q3: str = ""
    must: list[str] = []
    red_flags: list[str] = []

class UpdateScenarioRequest(BaseModel):
    name: Optional[str] = None
    business_type: Optional[str] = None
    shop_name: Optional[str] = None
    shop_category: Optional[str] = None
    shop_address: Optional[str] = None
    csv_name: Optional[str] = None
    q1: Optional[str] = None
    q2: Optional[str] = None
    q3: Optional[str] = None
    must: Optional[list[str]] = None
    red_flags: Optional[list[str]] = None
    is_active: Optional[bool] = None


def _scenario_to_dict(s: TestScenario) -> dict:
    shop = json.loads(s.shop_json) if s.shop_json else {}
    return {
        "id": s.id,
        "sid": s.sid,
        "name": s.name,
        "business_type": s.business_type,
        "shop_name": shop.get("name", ""),
        "shop_category": shop.get("category", ""),
        "shop_address": shop.get("address", ""),
        "csv_name": s.csv_name,
        "q1": s.q1,
        "q2": s.q2,
        "q3": s.q3,
        "must": json.loads(s.must_json) if s.must_json else [],
        "red_flags": json.loads(s.red_flags_json) if s.red_flags_json else [],
        "is_active": s.is_active,
        "created_at": s.created_at.isoformat(),
        "updated_at": s.updated_at.isoformat(),
    }


@router.get("/test-scenarios")
def list_test_scenarios(db: Session = Depends(get_db), admin=Depends(require_admin)):
    rows = db.exec(select(TestScenario).order_by(TestScenario.sid)).all()
    return [_scenario_to_dict(r) for r in rows]


@router.post("/test-scenarios")
def create_test_scenario(req: CreateScenarioRequest, db: Session = Depends(get_db), admin=Depends(require_admin)):
    existing = db.exec(select(TestScenario).where(TestScenario.sid == req.sid)).first()
    if existing:
        raise HTTPException(400, f"sid '{req.sid}' already exists")
    shop = {"name": req.shop_name, "category": req.shop_category, "address": req.shop_address}
    row = TestScenario(
        sid=req.sid,
        name=req.name,
        business_type=req.business_type,
        shop_json=json.dumps(shop, ensure_ascii=False),
        csv_name=req.csv_name,
        q1=req.q1,
        q2=req.q2,
        q3=req.q3,
        must_json=json.dumps(req.must, ensure_ascii=False),
        red_flags_json=json.dumps(req.red_flags, ensure_ascii=False),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _scenario_to_dict(row)


@router.patch("/test-scenarios/{scenario_id}")
def update_test_scenario(scenario_id: int, req: UpdateScenarioRequest,
                          db: Session = Depends(get_db), admin=Depends(require_admin)):
    row = db.get(TestScenario, scenario_id)
    if not row:
        raise HTTPException(404, "scenario not found")

    if req.name is not None:
        row.name = req.name
    if req.business_type is not None:
        row.business_type = req.business_type
    if req.csv_name is not None:
        row.csv_name = req.csv_name
    if req.q1 is not None:
        row.q1 = req.q1
    if req.q2 is not None:
        row.q2 = req.q2
    if req.q3 is not None:
        row.q3 = req.q3
    if req.must is not None:
        row.must_json = json.dumps(req.must, ensure_ascii=False)
    if req.red_flags is not None:
        row.red_flags_json = json.dumps(req.red_flags, ensure_ascii=False)
    if req.is_active is not None:
        row.is_active = req.is_active

    # Update shop_json if any shop field provided
    if req.shop_name is not None or req.shop_category is not None or req.shop_address is not None:
        shop = json.loads(row.shop_json) if row.shop_json else {}
        if req.shop_name is not None:
            shop["name"] = req.shop_name
        if req.shop_category is not None:
            shop["category"] = req.shop_category
        if req.shop_address is not None:
            shop["address"] = req.shop_address
        row.shop_json = json.dumps(shop, ensure_ascii=False)

    row.updated_at = datetime.now(timezone.utc)
    db.add(row)
    db.commit()
    db.refresh(row)
    return _scenario_to_dict(row)


@router.delete("/test-scenarios/{scenario_id}")
def delete_test_scenario(scenario_id: int, db: Session = Depends(get_db), admin=Depends(require_admin)):
    row = db.get(TestScenario, scenario_id)
    if not row:
        raise HTTPException(404, "scenario not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


# ── Bulk import ───────────────────────────────────────────────────────────────

_TEMPLATE_PATH = Path(__file__).parent / "scenario_import_template.xlsx"

_REQUIRED_COLS = {"sid", "name"}
_ALL_COLS = ["sid", "name", "business_type", "shop_name", "shop_category",
             "shop_address", "csv_name", "q1", "q2", "q3", "must", "red_flags"]


@router.get("/test-scenarios/bulk-import/template")
def download_bulk_import_template(admin=Depends(require_admin)):
    """Download the Excel template with instructions."""
    if not _TEMPLATE_PATH.exists():
        raise HTTPException(404, "Template file not found on server")
    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(_TEMPLATE_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="scenario_import_template.xlsx",
    )


@router.post("/test-scenarios/bulk-import")
async def bulk_import_scenarios(
    file: UploadFile = File(...),
    overwrite: bool = Form(False),
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Upload a zip containing:
      - one .xlsx file (scenarios) — must have a sheet named "场景数据"
      - optional data files named {sid}.csv or {sid}.xlsx

    Returns a summary: created / updated / skipped / errors.
    """
    import zipfile
    import tempfile
    import openpyxl

    if not file.filename.endswith(".zip"):
        raise HTTPException(400, "请上传 .zip 文件")

    content = await file.read()

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        zip_path = tmp / "upload.zip"
        zip_path.write_bytes(content)

        try:
            with zipfile.ZipFile(zip_path) as zf:
                zf.extractall(tmp)
        except zipfile.BadZipFile:
            raise HTTPException(400, "zip 文件损坏，请重新压缩后上传")

        # Find the Excel file (exclude macOS __MACOSX junk)
        xlsx_files = [
            p for p in tmp.rglob("*.xlsx")
            if "__MACOSX" not in str(p) and p.name != "upload.zip"
        ]
        if not xlsx_files:
            raise HTTPException(400, "zip 内未找到 .xlsx 文件")
        if len(xlsx_files) > 1:
            raise HTTPException(400, f"zip 内有多个 .xlsx 文件，请只保留一个场景表：{[f.name for f in xlsx_files]}")

        xlsx_path = xlsx_files[0]
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(422, f"无法读取 Excel 文件：{e}")

        if "场景数据" not in wb.sheetnames:
            raise HTTPException(422, f"Excel 中未找到「场景数据」sheet，现有 sheet：{wb.sheetnames}")

        ws = wb["场景数据"]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            raise HTTPException(422, "「场景数据」sheet 内容为空（至少需要表头行+说明行+一条数据）")

        # Row 1 = headers, Row 2 = hints (skip), Row 3+ = data
        raw_headers = [str(h).strip() if h else "" for h in rows[0]]
        col_idx = {h: i for i, h in enumerate(raw_headers) if h in _ALL_COLS}

        missing = _REQUIRED_COLS - set(col_idx.keys())
        if missing:
            raise HTTPException(422, f"Excel 缺少必填列：{missing}，请使用官方模板")

        def _get(row, col_name, default=""):
            idx = col_idx.get(col_name)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        # Collect data attachments from zip (flat — any depth)
        data_files = {
            p.name: p for p in tmp.rglob("*")
            if p.is_file()
            and p.suffix.lower() in (".csv", ".xlsx", ".xls")
            and "__MACOSX" not in str(p)
            and p != xlsx_path
        }

        SCENARIO_DATA_DIR.mkdir(parents=True, exist_ok=True)

        results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

        for row_num, row in enumerate(rows[2:], start=3):
            sid = _get(row, "sid")
            name = _get(row, "name")

            # Skip blank rows
            if not sid and not name:
                continue

            # Validate
            row_errors = []
            if not sid:
                row_errors.append("sid 为空")
            if not name:
                row_errors.append("name 为空")
            if not sid.replace("_", "").isalnum():
                row_errors.append(f"sid '{sid}' 含非法字符（只允许字母和数字）")
            if row_errors:
                results["errors"].append({"row": row_num, "sid": sid or "?", "reason": "；".join(row_errors)})
                continue

            csv_name = _get(row, "csv_name")

            # Copy data attachment if present in zip
            if csv_name:
                src = data_files.get(csv_name)
                if src:
                    dest = SCENARIO_DATA_DIR / csv_name
                    import shutil
                    shutil.copy2(src, dest)
                else:
                    results["errors"].append({
                        "row": row_num, "sid": sid,
                        "reason": f"csv_name='{csv_name}' 在 zip 内未找到对应文件",
                    })
                    # Don't abort — still import the scenario record

            # Parse must / red_flags (semicolon separated)
            def _split(val):
                return [x.strip() for x in val.split(";") if x.strip()]

            shop = {
                "name":     _get(row, "shop_name"),
                "category": _get(row, "shop_category"),
                "address":  _get(row, "shop_address"),
            }

            existing = db.exec(select(TestScenario).where(TestScenario.sid == sid)).first()
            if existing:
                if not overwrite:
                    results["skipped"] += 1
                    continue
                existing.name          = name
                existing.business_type = _get(row, "business_type")
                existing.shop_json     = json.dumps(shop, ensure_ascii=False)
                existing.csv_name      = csv_name
                existing.q1            = _get(row, "q1")
                existing.q2            = _get(row, "q2")
                existing.q3            = _get(row, "q3")
                existing.must_json     = json.dumps(_split(_get(row, "must")), ensure_ascii=False)
                existing.red_flags_json= json.dumps(_split(_get(row, "red_flags")), ensure_ascii=False)
                existing.updated_at    = datetime.now(timezone.utc)
                db.add(existing)
                results["updated"] += 1
            else:
                row_obj = TestScenario(
                    sid           = sid,
                    name          = name,
                    business_type = _get(row, "business_type"),
                    shop_json     = json.dumps(shop, ensure_ascii=False),
                    csv_name      = csv_name,
                    q1            = _get(row, "q1"),
                    q2            = _get(row, "q2"),
                    q3            = _get(row, "q3"),
                    must_json     = json.dumps(_split(_get(row, "must")), ensure_ascii=False),
                    red_flags_json= json.dumps(_split(_get(row, "red_flags")), ensure_ascii=False),
                )
                db.add(row_obj)
                results["created"] += 1

        db.commit()

    return {
        "ok": True,
        "created": results["created"],
        "updated": results["updated"],
        "skipped": results["skipped"],
        "errors":  results["errors"],
        "total_processed": results["created"] + results["updated"] + results["skipped"] + len(results["errors"]),
    }
